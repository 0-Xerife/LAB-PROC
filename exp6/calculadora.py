#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
PCS3732 - EXP6
Calculadora binária no Raspberry Pi 3 com complemento de 2.

Modos:

1) Modo normal:
   python3 calculadora_raspberry.py --mode normal --bits-width 4

2) Benchmark CSV:
   python3 calculadora_raspberry.py --mode benchmark --output results/benchmark_raspberry.csv

3) Benchmark XLSX:
   python3 calculadora_raspberry.py --mode benchmark --output results/benchmark_raspberry.xlsx

Características:
- Entradas interpretadas em complemento de 2.
- Operações com largura fixa de n bits.
- Resultado com overflow/wrap em n bits.
- Benchmark mede as operações em lote (múltiplas iterações por amostra) para diluir o overhead.
- O benchmark salva média, desvio padrão, mínimo e máximo por operação isolada.
"""

from __future__ import annotations

import argparse
import csv
import math
import statistics
import time
from pathlib import Path


# ============================================================
# Complemento de 2
# ============================================================

def signed_range(width: int) -> tuple[int, int]:
    if width < 2:
        raise ValueError("Complemento de 2 exige pelo menos 2 bits.")
    return -(1 << (width - 1)), (1 << (width - 1)) - 1


def parse_twos(binary_string: str, width: int) -> int:
    """
    Converte string binária para inteiro com sinal em complemento de 2.
    """
    s = binary_string.strip().replace(" ", "").replace("_", "")

    if not s:
        raise ValueError("entrada vazia")

    if any(c not in "01" for c in s):
        raise ValueError("a entrada deve conter apenas 0 e 1")

    if len(s) > width:
        raise ValueError(f"entrada possui {len(s)} bits, mas a largura é {width}")

    s = s.zfill(width)

    unsigned_value = int(s, 2)
    sign_bit = 1 << (width - 1)

    if unsigned_value & sign_bit:
        return unsigned_value - (1 << width)

    return unsigned_value


def to_twos(value: int, width: int) -> str:
    """
    Converte inteiro para binário em complemento de 2 com largura fixa.
    Aplica truncamento/wrap.
    """
    mask = (1 << width) - 1
    return format(value & mask, f"0{width}b")


def wrap_twos(value: int, width: int) -> int:
    """
    Aplica overflow de complemento de 2 para largura fixa.
    """
    mask = (1 << width) - 1
    value &= mask

    sign_bit = 1 << (width - 1)

    if value & sign_bit:
        return value - (1 << width)

    return value


# ============================================================
# Operações
# ============================================================

def calc_add(a: int, b: int, width: int) -> int:
    return wrap_twos(a + b, width)


def calc_sub(a: int, b: int, width: int) -> int:
    return wrap_twos(a - b, width)


def calc_mul(a: int, b: int, width: int) -> int:
    return wrap_twos(a * b, width)


def calc_div(a: int, b: int, width: int) -> int:
    if b == 0:
        raise ZeroDivisionError("divisão por zero")
    return wrap_twos(int(a / b), width)


def calc_factorial(a: int, width: int) -> int:
    if a < 0:
        raise ValueError("fatorial não definido para número negativo")
    return wrap_twos(math.factorial(a), width)


# ============================================================
# Modo normal
# ============================================================

def run_normal_mode(width: int) -> None:
    min_value, max_value = signed_range(width)

    print("=" * 72)
    print("Calculadora Binária - Raspberry Pi 3")
    print("Modo normal com complemento de 2")
    print("=" * 72)
    print(f"Largura fixa: {width} bits")
    print(f"Intervalo decimal representável: {min_value} até {max_value}")
    print()
    print("Operações:")
    print("  +  soma")
    print("  -  subtração")
    print("  * multiplicação")
    print("  /  divisão inteira")
    print("  !  fatorial")
    print("  q  sair")
    print()
    print("Exemplos em 4 bits:")
    print("  0111 = 7")
    print("  1000 = -8")
    print("  1111 = -1")
    print("=" * 72)
    print()

    while True:
        op = input("Operação (+, -, *, /, !, q): ").strip()

        if op.lower() in {"q", "quit", "sair"}:
            print("Encerrando.")
            break

        if op not in {"+", "-", "*", "/", "!"}:
            print("ERRO: operação inválida.\n")
            continue

        try:
            a_bin = input(f"A em binário, até {width} bits: ")
            a = parse_twos(a_bin, width)
            b = None

            if op != "!":
                b_bin = input(f"B em binário, até {width} bits: ")
                b = parse_twos(b_bin, width)

            if op == "+":
                result = calc_add(a, b, width)
                expression = f"{a} + {b}"
            elif op == "-":
                result = calc_sub(a, b, width)
                expression = f"{a} - {b}"
            elif op == "*":
                result = calc_mul(a, b, width)
                expression = f"{a} * {b}"
            elif op == "/":
                result = calc_div(a, b, width)
                expression = f"{a} / {b}"
            elif op == "!":
                result = calc_factorial(a, width)
                expression = f"{a}!"
            else:
                raise ValueError("operação inválida")

            print()
            print("Resultado:")
            print(f"Expressão decimal interpretada: {expression}")
            print(f"Resultado decimal em {width} bits: {result}")
            print(f"Resultado binário em complemento de 2: {to_twos(result, width)}")
            print()

        except ZeroDivisionError:
            print("\nERRO: divisão por zero. O programa não travou e aguarda nova entrada.\n")
        except Exception as exc:
            print(f"\nERRO: {exc}\n")


# ============================================================
# Benchmark
# ============================================================

def make_operand(bits: int, salt: int) -> int:
    """Gera operandos determinísticos dentro do intervalo de complemento de 2."""
    min_value, max_value = signed_range(bits)
    span = max_value - min_value + 1
    raw = (0x9E3779B97F4A7C15 * (salt + 1)) % span
    return min_value + raw


def measure_operation_batch(op: str, bits: int, iterations: int) -> tuple[float, int, int, int]:
    """
    Mede a operação rodando um lote (batch) de N iterações e divide o tempo total.
    Isso dilui o overhead do timer e do escalonador do SO.
    """
    a = make_operand(bits, salt=1)
    b = make_operand(bits, salt=2)
    result = 0

    if op == "!":
        _, max_value = signed_range(bits)
        a = min(bits, max_value)
        b = None

        start = time.perf_counter_ns()
        for _ in range(iterations):
            result = calc_factorial(a, bits)
        end = time.perf_counter_ns()

    elif op == "+":
        start = time.perf_counter_ns()
        for _ in range(iterations):
            result = calc_add(a, b, bits)
        end = time.perf_counter_ns()

    elif op == "-":
        start = time.perf_counter_ns()
        for _ in range(iterations):
            result = calc_sub(a, b, bits)
        end = time.perf_counter_ns()

    elif op == "*":
        start = time.perf_counter_ns()
        for _ in range(iterations):
            result = calc_mul(a, b, bits)
        end = time.perf_counter_ns()

    else:
        raise ValueError(f"operação inválida no benchmark: {op}")

    # Retorna o tempo médio por UMA operação em nanosegundos
    elapsed_ns_per_op = (end - start) / iterations
    return elapsed_ns_per_op, a, b, result


def operation_name(op: str) -> str:
    return {
        "+": "soma",
        "-": "subtracao",
        "*": "multiplicacao",
        "!": "fatorial",
    }[op]


def run_benchmark(bits_list: list[int], samples: int, iterations: int) -> list[dict]:
    operations = ["+", "-", "*", "!"]
    results = []

    for bits in bits_list:
        min_value, max_value = signed_range(bits)

        for op in operations:
            measurements = []
            example_a = None
            example_b = None
            example_result = None

            for _ in range(samples):
                elapsed_ns, a, b, result = measure_operation_batch(op, bits, iterations)
                measurements.append(elapsed_ns)

                example_a = a
                example_b = b
                example_result = result

            mean_ns = statistics.mean(measurements)
            stdev_ns = statistics.stdev(measurements) if samples > 1 else 0.0
            min_ns = min(measurements)
            max_ns = max(measurements)

            row = {
                "plataforma": "Raspberry Pi 3",
                "linguagem": "Python 3",
                "representacao": "complemento de 2",
                "bits": bits,
                "intervalo_decimal": f"{min_value} a {max_value}",
                "operacao": operation_name(op),
                "simbolo": op,
                "amostras": samples,
                "iteracoes_por_amostra": iterations,
                "tempo_medio_ns": mean_ns,
                "desvio_padrao_ns": stdev_ns,
                "tempo_minimo_ns": min_ns,
                "tempo_maximo_ns": max_ns,
                "tempo_medio_us": mean_ns / 1000,
                "A_decimal": example_a,
                "A_binario": to_twos(example_a, bits),
                "B_decimal": "" if example_b is None else example_b,
                "B_binario": "" if example_b is None else to_twos(example_b, bits),
                "resultado_decimal": example_result,
                "resultado_binario": to_twos(example_result, bits),
                "observacao": (
                    f"cada amostra computou a media de {iterations} execucoes em lote para estabilidade"
                ),
            }

            results.append(row)

            print(
                f"bits={bits:>3} | "
                f"op={operation_name(op):<14} | "
                f"média={mean_ns:>10.2f} ns | "
                f"desvio={stdev_ns:>10.2f} ns | "
                f"min={min_ns:>8.0f} ns | "
                f"max={max_ns:>8.0f} ns"
            )

    return results


# ============================================================
# Salvar CSV / XLSX
# ============================================================

FIELDS = [
    "plataforma",
    "linguagem",
    "representacao",
    "bits",
    "intervalo_decimal",
    "operacao",
    "simbolo",
    "amostras",
    "iteracoes_por_amostra",
    "tempo_medio_ns",
    "desvio_padrao_ns",
    "tempo_minimo_ns",
    "tempo_maximo_ns",
    "tempo_medio_us",
    "A_decimal",
    "A_binario",
    "B_decimal",
    "B_binario",
    "resultado_decimal",
    "resultado_binario",
    "observacao",
]


def save_csv(results: list[dict], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=FIELDS)
        writer.writeheader()
        writer.writerows(results)


def save_xlsx(results: list[dict], output_path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError(
            "Para salvar XLSX, instale openpyxl com: pip3 install openpyxl"
        ) from exc

    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Benchmark RPi3"

    ws.append(FIELDS)

    for row in results:
        ws.append([row[field] for field in FIELDS])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for col_idx, field in enumerate(FIELDS, start=1):
        max_len = len(field)
        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 45)

    ws.freeze_panes = "A2"
    wb.save(output_path)


def save_results(results: list[dict], output: str) -> None:
    output_path = Path(output)

    if output_path.suffix.lower() == ".csv":
        save_csv(results, output_path)
    elif output_path.suffix.lower() == ".xlsx":
        save_xlsx(results, output_path)
    else:
        raise ValueError("formato inválido. Use .csv ou .xlsx")

    print()
    print(f"Resultados salvos em: {output_path.resolve()}")


# ============================================================
# CLI
# ============================================================

def parse_bits_list(bits_string: str) -> list[int]:
    bits_list = []
    for item in bits_string.split(","):
        item = item.strip()
        if not item:
            continue
        value = int(item)
        if value < 2:
            raise ValueError("use pelo menos 2 bits para complemento de 2")
        bits_list.append(value)

    if not bits_list:
        raise ValueError("lista de bits vazia")
    return bits_list


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Calculadora binária no Raspberry Pi 3 com complemento de 2"
    )

    parser.add_argument(
        "--mode",
        choices=["normal", "benchmark"],
        default="normal",
        help="modo de execução",
    )

    parser.add_argument(
        "--bits-width",
        type=int,
        default=4,
        help="largura fixa em bits no modo normal",
    )

    parser.add_argument(
        "--bits",
        default="4,8,16,32,64",
        help="larguras para benchmark separadas por vírgula",
    )

    parser.add_argument(
        "--samples",
        type=int,
        default=30,
        help="número de amostras capturadas",
    )

    parser.add_argument(
        "--iterations",
        type=int,
        default=5000,
        help="número de execuções agrupadas por amostra (estabiliza o tempo)",
    )

    parser.add_argument(
        "--output",
        default="results/benchmark_raspberry_c2.csv",
        help="arquivo de saída .csv ou .xlsx",
    )

    args = parser.parse_args()

    if args.bits_width < 2:
        raise ValueError("bits-width deve ser pelo menos 2")

    if args.mode == "normal":
        run_normal_mode(width=args.bits_width)

    elif args.mode == "benchmark":
        bits_list = parse_bits_list(args.bits)

        print("=" * 72)
        print("Benchmark - Raspberry Pi 3")
        print("Representação: complemento de 2")
        print(f"Cada amostra é a média de um lote de {args.iterations} execuções consecutivas.")
        print("=" * 72)
        print(f"Bits testados: {bits_list}")
        print(f"Amostras geradas: {args.samples}")
        print(f"Iterações por amostra: {args.iterations}")
        print(f"Arquivo de saída: {args.output}")
        print("=" * 72)
        print()

        results = run_benchmark(
            bits_list=bits_list,
            samples=args.samples,
            iterations=args.iterations
        )

        save_results(results, args.output)


if __name__ == "__main__":
    main()


